#!/usr/bin/env python3
"""
Generate customized cover letters for all 605 job opportunities.
Organizes by fit tier (EXCELLENT, STRONG, GOOD, POSSIBLE) for easy navigation.
"""

import re
import openpyxl
from pathlib import Path
from collections import defaultdict

# Company-specific profiles
COMPANY_PROFILES = {
    "Wise": {
        "custom_intro": "Over the past five years at PSP Investments (~$300B CAD AUM), I have designed, built, and deployed production-grade data infrastructure systems that power institutional investment operations. At Wise, I see an opportunity to bring deep experience in building scalable data platforms and real-time analytics infrastructure to support fintech at global scale.",
        "project_highlight": "The highlight of my work has been leading end-to-end data infrastructure projects, including building custom data pipelines on Databricks (Spark, Delta Lake) that process millions of records daily, supporting analytics, ML, and AI workloads. I designed comprehensive financial and alternative data pipelines with point-in-time accuracy, built distributed systems handling high-throughput transactions, and deployed production ML pipelines with strict SLA requirements. Every system I build prioritizes reliability, scalability, and operational clarity—critical for fintech.",
        "tech_skills": "Python, Databricks, Apache Spark, Delta Lake, SQL, Azure, Kafka, distributed systems design, and data platform architecture",
        "alignment_paragraph": "Wise's mission to make international money movement seamless aligns with my passion for building robust financial infrastructure. I bring experience building systems that handle billions in transactions reliably, with deep expertise in data platform design, real-time processing, and the operational discipline required for fintech at scale. I am excited to bring this blend of infrastructure engineering, fintech domain knowledge, and production deployment experience to your team."
    },
    "Anthropic": {
        "custom_intro": "Over the past five years, I have designed and deployed production-grade agentic AI systems at scale. I led development of PSP's Virtual Analyst Platform—a flagship agentic AI system serving 300+ institutional users—and own the full ML infrastructure stack from LLM integrations to advanced retrieval systems. At Anthropic, I see the opportunity to apply this expertise to advance AI infrastructure at the frontier.",
        "project_highlight": "The highlight of my work has been architecting and deploying agentic AI systems that institutional teams trust and use daily. I built the underlying infrastructure: custom SDK with vector search retrieval, multi-step reasoning, function calling, and integration of multiple LLM providers (OpenAI, Gemini, Databricks). I designed systems serving 100+ active users with mission-critical use cases, ensuring reliability and interpretability in every deployment. I have hands-on expertise in AI infrastructure, LLM optimization, and building systems that users trust.",
        "tech_skills": "Python, LangChain, LiteLLM, LLM APIs (OpenAI, Gemini), Azure AI Foundry, Gradio, Databricks, vector databases, and agentic AI architecture",
        "alignment_paragraph": "Anthropic's mission to develop safe, interpretable AI systems resonates deeply with my work on transparent agentic systems. I bring hands-on experience building AI infrastructure that serves real users, deep expertise in LLM integrations and multi-model reasoning, and a track record of shipping production systems reliably. I am excited to contribute to advancing AI infrastructure at Anthropic."
    },
    "Perplexity AI": {
        "custom_intro": "Over the past five years, I have built production-grade AI infrastructure systems that power institutional research and decision-making. I designed and deployed advanced retrieval systems, built multi-step reasoning pipelines, and led infrastructure for search-driven AI at scale. At Perplexity, I see the opportunity to bring deep expertise in search, retrieval, and agentic systems to revolutionize how people access knowledge.",
        "project_highlight": "The highlight of my work has been architecting retrieval and reasoning systems for institutional AI. I built custom vector search pipelines across thousands of documents with source traceability, implemented multi-step reasoning workflows for complex queries, and deployed systems serving 100+ active users. I have deep hands-on experience with retrieval-augmented generation (RAG), semantic search, and ensuring every answer links back to its source—critical for trustworthy search and discovery.",
        "tech_skills": "Python, vector databases, semantic search, LLM APIs, retrieval-augmented generation (RAG), Databricks, SQL, and information retrieval systems",
        "alignment_paragraph": "Perplexity's approach to building intelligent search systems grounded in verifiable sources aligns perfectly with my expertise in building trustworthy retrieval infrastructure. I bring hands-on experience with RAG systems, vector search optimization, and the full stack from data ingestion to search-driven AI interfaces. I am excited to apply this expertise to advancing search and discovery at Perplexity."
    },
    "Pinecone": {
        "custom_intro": "Over the past five years, I have built production AI infrastructure systems centered on vector databases and semantic search. I designed and deployed retrieval pipelines for institutional AI, integrated vector search into complex agentic systems, and led architecture for retrieval-augmented generation at scale. At Pinecone, I see the opportunity to bring deep expertise in vector database usage and production RAG systems.",
        "project_highlight": "The highlight of my work has been architecting production retrieval systems powered by vector search and semantic similarity. I built custom pipelines that index thousands of financial documents, implemented multi-step retrieval workflows for complex queries, and deployed systems serving institutional teams daily. I have hands-on expertise integrating vector databases into production systems, optimizing retrieval quality and latency, and designing search-driven AI applications.",
        "tech_skills": "Python, vector databases, semantic search, embeddings, retrieval-augmented generation (RAG), LLM integrations, and database optimization",
        "alignment_paragraph": "Pinecone's vision to democratize vector search and build the foundational infrastructure for AI applications aligns with my passion for building scalable retrieval systems. I bring hands-on experience building production RAG systems at scale, deep expertise in vector database integration and optimization, and a track record of shipping systems that developers trust. I am excited to contribute to advancing vector database infrastructure at Pinecone."
    },
    "Confluent": {
        "custom_intro": "Over the past five years, I have designed and built production data infrastructure systems at scale, including real-time data pipelines, event streaming architectures, and distributed data platforms. I designed systems handling millions of events daily for institutional investment operations. At Confluent, I see the opportunity to bring deep expertise in event streaming infrastructure and data platform architecture.",
        "project_highlight": "The highlight of my work has been architecting production data platforms with real-time pipelines and reliable event streaming. I built comprehensive data pipelines on Databricks processing millions of records daily, designed data architecture for institutional-scale operations, and implemented systems with strict reliability and latency requirements. I have hands-on expertise with distributed data systems, stream processing, and building infrastructure that serves critical business operations.",
        "tech_skills": "Python, Kafka, Apache Spark, Delta Lake, Databricks, SQL, event streaming architecture, and distributed data systems",
        "alignment_paragraph": "Confluent's mission to build real-time data infrastructure resonates with my expertise in building production data pipelines at scale. I bring hands-on experience designing event streaming systems, deep expertise in data architecture and platform engineering, and a track record of shipping systems that handle mission-critical operations reliably. I am excited to bring this expertise to advancing event streaming infrastructure at Confluent."
    },
    "Stripe": {
        "custom_intro": "Over the past five years at PSP Investments, I have built production data infrastructure powering institutional operations and decision-making. I designed comprehensive financial data pipelines, built ML systems for predictive analytics, and led architecture for real-time data serving billions in transactions reliably. At Stripe, I see the opportunity to apply this fintech expertise and data infrastructure experience at global scale.",
        "project_highlight": "The highlight of my work has been designing production data systems for financial operations, including building pipelines for financial and alternative data with point-in-time accuracy, designing ML systems for institutional prediction tasks, and architecting platforms serving 100+ active users. I have deep expertise in financial data modeling, payment systems understanding, and building infrastructure that handles high-throughput, low-latency requirements critical for fintech.",
        "tech_skills": "Python, Databricks, Apache Spark, SQL, financial data modeling, ML infrastructure, and real-time data systems",
        "alignment_paragraph": "Stripe's mission to increase the GDP of the internet through robust payment infrastructure aligns with my passion for building reliable fintech systems. I bring hands-on experience in financial data infrastructure, deep expertise in building systems that handle mission-critical transactions reliably, and a strong fintech domain background. I am excited to apply this expertise to advancing payment infrastructure at Stripe."
    },
    "xAI": {
        "custom_intro": "Over the past five years, I have designed and deployed production-grade ML infrastructure systems and agentic AI applications at scale. I led development of PSP's Virtual Analyst Platform and own the full ML infrastructure stack from LLM integrations to advanced reasoning systems. At xAI, I see the opportunity to bring deep expertise in AI infrastructure and reasoning systems.",
        "project_highlight": "The highlight of my work has been architecting production agentic AI systems with transparency and reliability. I built custom AI infrastructure: SDK with advanced retrieval, multi-step reasoning, function calling, and integration of multiple LLM providers. I designed systems serving 100+ active users with mission-critical use cases, ensuring reliability and interpretability in every deployment. I have hands-on expertise in AI infrastructure, LLM optimization, and building systems that users trust.",
        "tech_skills": "Python, LangChain, LiteLLM, LLM APIs, reasoning systems, Databricks, vector search, and agentic AI architecture",
        "alignment_paragraph": "xAI's mission to develop AI systems with deep understanding of physical reality aligns with my expertise in building production AI infrastructure. I bring hands-on experience building reasoning-heavy AI systems, deep expertise in LLM architecture and optimization, and a track record of shipping systems that users trust in mission-critical environments. I am excited to contribute to advancing AI reasoning at xAI."
    },
    "Cohere": {
        "custom_intro": "Over the past five years, I have designed and deployed production-grade AI infrastructure systems leveraging large language models at scale. I led development of agentic AI systems, built LLM-powered retrieval and reasoning platforms, and own expertise across the full ML infrastructure stack. At Cohere, I see the opportunity to bring deep experience in LLM infrastructure and AI systems development.",
        "project_highlight": "The highlight of my work has been architecting production LLM-driven systems serving institutional users. I built custom AI infrastructure integrating multiple LLM providers, implemented advanced retrieval and reasoning workflows, and deployed systems serving 100+ active users daily. I have hands-on expertise in LLM API optimization, multi-model reasoning, and building production systems that leverage cutting-edge language models reliably.",
        "tech_skills": "Python, LLM APIs, LangChain, semantic search, retrieval systems, Databricks, and AI infrastructure",
        "alignment_paragraph": "Cohere's mission to democratize access to advanced language models aligns with my passion for building production AI infrastructure. I bring hands-on experience integrating and optimizing LLM systems at scale, deep expertise in building AI applications developers trust, and a track record of shipping systems that serve real users in demanding environments. I am excited to contribute to advancing language models at Cohere."
    },
    "Harvey AI": {
        "custom_intro": "Over the past five years, I have designed and deployed production-grade AI infrastructure systems that serve specialized professional domains. I built agentic AI platforms for institutional decision-making, implemented domain-specific reasoning systems, and led architecture for AI applications in regulated environments. At Harvey AI, I see the opportunity to bring deep expertise in building domain-specialized AI systems.",
        "project_highlight": "The highlight of my work has been architecting AI systems for professional domains with strict reliability and interpretability requirements. I built custom AI infrastructure with full transparency into reasoning, implemented specialized retrieval systems, and deployed to users in mission-critical environments. I have hands-on expertise in building AI systems that professionals trust, ensuring every decision is grounded in verifiable sources and clear reasoning.",
        "tech_skills": "Python, LLM APIs, domain-specific reasoning, retrieval systems, Databricks, semantic search, and professional AI systems",
        "alignment_paragraph": "Harvey AI's mission to augment professional expertise with AI aligns with my expertise in building trustworthy AI systems for specialized domains. I bring hands-on experience building AI platforms for professionals, deep expertise in reasoning system design and transparent decision-making, and a track record of shipping systems in regulated environments. I am excited to contribute to advancing AI for professional services at Harvey AI."
    },
    "Generic": {
        "custom_intro": "Over the past five years at PSP Investments (~$300B CAD AUM), I have designed, built, and deployed production-grade infrastructure systems that serve institutional operations. I bring deep expertise in building scalable systems from prototype to live production.",
        "project_highlight": "The highlight of my work has been leading end-to-end infrastructure projects including designing comprehensive data pipelines, building distributed systems, and architecting platforms serving 100+ active users. I have hands-on expertise in the full stack: backend architecture, system design, integration of complex tools and systems, custom interface development, and production deployment.",
        "tech_skills": "Python, Databricks, Apache Spark, SQL, Azure, distributed systems design, LLM APIs, and platform architecture",
        "alignment_paragraph": "I am excited to bring this blend of infrastructure engineering expertise, production deployment know-how, and experience building systems at scale to your team. I thrive in ambiguous, high-stakes environments and am quick to learn and ship new tools. I am an EU citizen with right to work in the EU; no sponsorship required. I am relocating to Europe and remote-ready."
    }
}

COVER_LETTER_TEMPLATE = """Dear Hiring Manager,

{custom_intro}

{project_highlight}

I bring strong proficiency in {tech_skills}, along with a B.Com in Finance & Business Analytics from McGill University (Desautels) and a DEC in Computer Science & Mathematics from Champlain College (Honor Roll). I hold Microsoft Azure AI Fundamentals certification. Fluent in English, French, and Romanian, with basic Spanish and Italian, I thrive in ambiguous, high-stakes environments and am quick to learn and ship new tools. I am an EU citizen with right to work in the EU; no sponsorship required. I am relocating to Europe and remote-ready.

{alignment_paragraph}

I would welcome the opportunity to discuss how I can contribute. I am available at your earliest convenience and can be reached at +1.514.710.9601 or chris.stefan@proton.me. You can also view my portfolio at chrizefan.github.io/me and digithings.ai.

Thank you for your consideration. I look forward to hearing from you.

Sincerely,
Chris Stefan"""

def sanitize_filename(text):
    """Remove invalid filename characters from text."""
    text = re.sub(r'[\[\]():/<>?*"|\\]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    text = text.replace(' ', '_')
    text = re.sub(r'[^\w_-]', '', text)
    return text

def get_company_profile(company_name):
    """Get the profile for a company, or return generic if not found."""
    if company_name in COMPANY_PROFILES:
        return COMPANY_PROFILES[company_name]

    company_lower = company_name.lower()
    for key in COMPANY_PROFILES.keys():
        if key.lower() in company_lower or company_lower in key.lower():
            return COMPANY_PROFILES[key]

    return COMPANY_PROFILES["Generic"]

def determine_tier(fit_score):
    """Determine fit tier based on score."""
    if fit_score >= 9.0:
        return "EXCELLENT"
    elif fit_score >= 8.0:
        return "STRONG"
    elif fit_score >= 7.0:
        return "GOOD"
    else:
        return "POSSIBLE"

def generate_all_cover_letters():
    """Generate cover letters for all 605 job opportunities."""

    print("Loading original cover letter template...")
    with open('/sessions/gifted-charming-mayer/mnt/me/cover_letter.txt', 'r') as f:
        original = f.read()
    print(f"✓ Original loaded ({len(original)} chars)")

    # Load cleaned jobs
    print("Loading cleaned job opportunities...")
    wb = openpyxl.load_workbook('/sessions/gifted-charming-mayer/mnt/me/Job_Opportunities_CLEANED.xlsx')
    ws = wb.active

    total_rows = ws.max_row - 1
    print(f"✓ Loaded {total_rows} job opportunities")

    # Create output structure by tier
    base_dir = Path('/sessions/gifted-charming-mayer/mnt/me/Cover_Letters')
    tiers = {}
    for tier in ["EXCELLENT", "STRONG", "GOOD", "POSSIBLE"]:
        tier_dir = base_dir / tier
        tier_dir.mkdir(parents=True, exist_ok=True)
        tiers[tier] = tier_dir

    # Track statistics
    stats = defaultdict(int)
    tier_counts = defaultdict(int)
    generated = []
    errors = []

    print(f"\nGenerating {total_rows} cover letters organized by fit tier...")
    print("=" * 80)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        try:
            company = row[1]  # Column B (Company)
            role = row[2]     # Column C (Job Title)
            fit_score = row[6]  # Column G (Fit Score Recalibrated - numeric)
            tier = row[7]     # Column H (Match Strength - tier name)

            if not company or not tier:
                continue

            # Use tier directly from spreadsheet
            tier_dir = tiers[tier]
            tier_counts[tier] += 1

            # Create filename
            company_clean = sanitize_filename(company or "Unknown")
            if not company_clean:
                company_clean = f"Role_{idx}"

            filename = f"{idx:04d}_{company_clean}.txt"
            filepath = tier_dir / filename

            # Skip if exists
            if filepath.exists():
                stats['skipped'] += 1
                continue

            # Get profile and generate letter
            profile = get_company_profile(company)
            letter = COVER_LETTER_TEMPLATE.format(
                custom_intro=profile["custom_intro"],
                project_highlight=profile["project_highlight"],
                tech_skills=profile["tech_skills"],
                alignment_paragraph=profile["alignment_paragraph"]
            )

            # Write file
            with open(filepath, 'w') as f:
                f.write(letter)

            stats['generated'] += 1
            generated.append({
                'idx': idx,
                'company': company,
                'tier': tier,
                'fit_score': fit_score,
                'filename': filename
            })

            # Print progress every 50
            if idx % 50 == 0:
                print(f"  {idx:4d}/{total_rows} generated | {tier} tier")

        except Exception as e:
            errors.append({
                'idx': idx,
                'company': company,
                'error': str(e)
            })
            stats['errors'] += 1

    print("=" * 80)
    print(f"\nGeneration Summary:")
    print(f"  Total rows processed: {total_rows}")
    print(f"  Generated: {stats['generated']}")
    print(f"  Skipped (already exist): {stats['skipped']}")
    print(f"  Errors: {stats['errors']}")
    print(f"\nBy Tier:")
    for tier in ["EXCELLENT", "STRONG", "GOOD", "POSSIBLE"]:
        print(f"  {tier:<12}: {tier_counts[tier]:4d} letters")

    if errors:
        print(f"\nErrors encountered ({len(errors)}):")
        for err in errors[:5]:
            print(f"  - Row {err['idx']}: {err['error']}")

    # Create comprehensive tracking spreadsheet
    print(f"\nCreating comprehensive tracking spreadsheet...")
    create_tracking_spreadsheet(generated, tier_counts)

    print(f"\n✓ All cover letters saved to: {base_dir}")
    print(f"  Organized by tier: EXCELLENT, STRONG, GOOD, POSSIBLE")

def create_tracking_spreadsheet(generated, tier_counts):
    """Create a comprehensive tracking spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Summary sheet
    summary = wb.active
    summary.title = "Summary"

    summary['A1'] = "Cover Letters Generation Summary"
    summary['A1'].font = Font(bold=True, size=14)

    row = 3
    summary[f'A{row}'] = "Tier"
    summary[f'B{row}'] = "Count"
    summary[f'A{row}'].font = Font(bold=True)
    summary[f'B{row}'].font = Font(bold=True)

    for i, tier in enumerate(["EXCELLENT", "STRONG", "GOOD", "POSSIBLE"], 1):
        summary[f'A{row+i}'] = tier
        summary[f'B{row+i}'] = tier_counts.get(tier, 0)

    summary[f'A{row+5}'] = "Total"
    summary[f'A{row+5}'].font = Font(bold=True)
    summary[f'B{row+5}'] = f'=SUM(B{row+1}:B{row+4})'
    summary[f'B{row+5}'].font = Font(bold=True)

    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 15

    # Save
    output_path = '/sessions/gifted-charming-mayer/mnt/me/All_Cover_Letters_Tracking.xlsx'
    wb.save(output_path)
    print(f"✓ Tracking file created: All_Cover_Letters_Tracking.xlsx")

if __name__ == "__main__":
    generate_all_cover_letters()
