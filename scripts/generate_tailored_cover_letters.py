#!/usr/bin/env python3
"""
Generate individually tailored cover letters for all 605 job postings.
Each letter references the specific role, company, and key skills.
"""

import re
import openpyxl
from pathlib import Path
from collections import defaultdict

# Company-specific intro patterns
COMPANY_INTROS = {
    "Wise": "At Wise, I see an opportunity to bring deep experience in building scalable data platforms and real-time analytics infrastructure to support fintech at global scale.",
    "Anthropic": "At Anthropic, I see the opportunity to apply this expertise to advance AI infrastructure at the frontier.",
    "Perplexity AI": "At Perplexity, I see the opportunity to bring deep expertise in search, retrieval, and agentic systems to revolutionize how people access knowledge.",
    "Pinecone": "At Pinecone, I see the opportunity to bring deep expertise in vector database usage and production RAG systems.",
    "Confluent": "At Confluent, I see the opportunity to bring deep expertise in event streaming infrastructure and data platform architecture.",
    "Stripe": "At Stripe, I see the opportunity to apply this fintech expertise and data infrastructure experience at global scale.",
    "xAI": "At xAI, I see the opportunity to bring deep expertise in AI infrastructure and reasoning systems.",
    "Cohere": "At Cohere, I see the opportunity to bring deep experience in LLM infrastructure and AI systems development.",
    "Harvey AI": "At Harvey AI, I see the opportunity to bring deep expertise in building domain-specialized AI systems.",
}

def sanitize_filename(text):
    """Remove invalid filename characters from text."""
    text = re.sub(r'[\[\]():/<>?*"|\\]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    text = text.replace(' ', '_')
    text = re.sub(r'[^\w_-]', '', text)
    return text

def extract_skills_from_string(skills_str):
    """Extract individual skills from comma-separated string."""
    if not skills_str:
        return []
    return [s.strip() for s in str(skills_str).split(',') if s.strip()]

def get_company_intro(company_name):
    """Get company-specific intro, or generic if not found."""
    for key in COMPANY_INTROS.keys():
        if key.lower() in str(company_name).lower():
            return COMPANY_INTROS[key]
    return f"I see this opportunity at {company_name} as a chance to bring my expertise in building scalable, production-grade infrastructure systems to a team pushing the boundaries of what's possible."

def generate_tailored_letter(company, role, skills, company_intro_suffix):
    """Generate a cover letter tailored to the specific job posting."""

    # Parse skills
    skills_list = extract_skills_from_string(skills)
    skills_str = ", ".join(skills_list[:5]) if skills_list else "infrastructure systems and AI platforms"

    # Build role-specific references
    role_lower = role.lower() if role else ""
    is_ml_role = any(term in role_lower for term in ["ml", "machine learning", "ai", "llm", "inference"])
    is_data_role = any(term in role_lower for term in ["data", "pipeline", "analytics", "etl"])
    is_infra_role = any(term in role_lower for term in ["infrastructure", "platform", "backend", "systems"])

    # Tailored intro paragraph
    intro = f"Over the past five years at PSP Investments (~$300B CAD AUM), I have designed, built, and deployed production-grade infrastructure systems that power institutional operations. {company_intro_suffix}"

    # Role-specific project highlight
    if is_ml_role:
        project = f"For the {role} role at {company}, my most relevant experience comes from architecting and deploying production ML infrastructure systems serving institutional users. I built custom LLM-powered platforms with advanced retrieval and reasoning capabilities, implemented vector search pipelines, and designed systems serving 100+ active users in mission-critical environments. Every system I built prioritizes reliability, scalability, and interpretability—critical for {company}'s success in this space."
    elif is_data_role:
        project = f"For the {role} role at {company}, my most relevant experience comes from designing and building production data infrastructure at scale. I built comprehensive data pipelines on Databricks processing millions of records daily, designed systems handling high-throughput transactions, and implemented real-time data serving with strict SLA requirements. I bring hands-on expertise in exactly the tech stack {company} uses: building systems that are reliable, scalable, and operationally clear."
    else:
        project = f"For the {role} role at {company}, my most relevant experience comes from architecting production infrastructure systems. I've led end-to-end projects including designing comprehensive data pipelines, building distributed systems, and implementing platforms serving 100+ active users. I have hands-on expertise in the full stack: system architecture, data infrastructure, LLM integration, custom interface development, and production deployment—everything needed to succeed in this role."

    # Tailored tech skills (prioritize posted skills)
    if skills_list:
        priority_skills = ", ".join(skills_list[:3])
        other_skills = "Python, Databricks, Apache Spark, SQL, Azure, and distributed systems design"
        tech_skills = f"{priority_skills}, {other_skills}"
    else:
        tech_skills = "Python, Databricks, Apache Spark, SQL, Azure, and distributed systems design"

    # Tailored alignment paragraph
    if is_ml_role:
        alignment = f"{company}'s focus on advancing ML infrastructure aligns perfectly with my expertise in building production AI systems at scale. I bring hands-on experience building LLM-powered systems that users trust, deep expertise in {priority_skills if skills_list else 'reasoning systems and optimization'}, and a track record of shipping systems in demanding environments. I'm excited to bring this expertise to advancing {company}'s mission."
    elif is_data_role:
        alignment = f"{company}'s mission to build robust data infrastructure resonates with my expertise. I bring hands-on experience designing production data systems, deep expertise in {priority_skills if skills_list else 'pipeline architecture and data modeling'}, and a track record of delivering systems that handle mission-critical operations reliably. I'm excited to contribute to {company}'s platform."
    else:
        alignment = f"I'm excited to bring my blend of infrastructure engineering expertise, production deployment know-how, and experience building systems at scale to {company}. I thrive in ambiguous, high-stakes environments and am quick to learn and ship new tools. I'm an EU citizen with right to work in the EU; no sponsorship required."

    # Assemble full letter
    letter = f"""Dear Hiring Manager,

{intro}

{project}

I bring strong proficiency in {tech_skills}, along with a B.Com in Finance & Business Analytics from McGill University (Desautels) and a DEC in Computer Science & Mathematics from Champlain College (Honor Roll). I hold Microsoft Azure AI Fundamentals certification. Fluent in English, French, and Romanian, with basic Spanish and Italian, I thrive in ambiguous, high-stakes environments and am quick to learn and ship new tools. I am an EU citizen with right to work in the EU; no sponsorship required. I am relocating to Europe and remote-ready.

{alignment}

I would welcome the opportunity to discuss how I can contribute to {company}. I am available at your earliest convenience and can be reached at +1.514.710.9601 or chris.stefan@proton.me. You can also view my portfolio at chrizefan.github.io/me and digithings.ai.

Thank you for your consideration. I look forward to hearing from you.

Sincerely,
Chris Stefan"""

    return letter

def generate_all_tailored_letters():
    """Generate individually tailored cover letters for all 605 job opportunities."""

    print("Loading job opportunities...")
    wb = openpyxl.load_workbook('/sessions/gifted-charming-mayer/mnt/me/Job_Opportunities_CLEANED.xlsx')
    ws = wb.active

    total_rows = ws.max_row - 1
    print(f"✓ Loaded {total_rows} opportunities")

    # Create output structure by tier
    base_dir = Path('/sessions/gifted-charming-mayer/mnt/me/Cover_Letters')

    tiers = {}
    for tier in ["EXCELLENT", "STRONG", "GOOD", "POSSIBLE"]:
        tier_dir = base_dir / tier
        tier_dir.mkdir(parents=True, exist_ok=True)
        tiers[tier] = tier_dir

    # Track statistics
    tier_counts = defaultdict(int)
    stats = defaultdict(int)
    errors = []

    print(f"\nGenerating {total_rows} tailored cover letters...")
    print("=" * 80)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        try:
            company = row[1]      # Column B (Company)
            role = row[2]         # Column C (Job Title)
            skills = row[8]       # Column I (Key Skills)
            tier = row[7]         # Column H (Match Strength)

            if not company or not tier:
                continue

            # Get company intro suffix
            company_intro = get_company_intro(company)

            # Generate tailored letter
            letter = generate_tailored_letter(company, role, skills, company_intro)

            # Save to appropriate tier directory
            tier_dir = tiers[tier]
            tier_counts[tier] += 1

            company_clean = sanitize_filename(company or "Unknown")
            if not company_clean:
                company_clean = f"Role_{idx}"

            filename = f"{idx:04d}_{company_clean}.txt"
            filepath = tier_dir / filename

            with open(filepath, 'w') as f:
                f.write(letter)

            stats['generated'] += 1

            # Print progress
            if idx % 50 == 0:
                print(f"  {idx:4d}/{total_rows} tailored | {tier} tier")

        except Exception as e:
            errors.append({'idx': idx, 'company': company, 'error': str(e)})
            stats['errors'] += 1

    print("=" * 80)
    print(f"\nGeneration Summary:")
    print(f"  Total processed: {total_rows}")
    print(f"  Generated: {stats['generated']}")
    print(f"  Errors: {stats['errors']}")
    print(f"\nBy Tier (Tailored):")
    for tier in ["EXCELLENT", "STRONG", "GOOD", "POSSIBLE"]:
        print(f"  {tier:<12}: {tier_counts[tier]:4d} letters")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:5]:
            print(f"  - Row {err['idx']}: {err['error']}")

    print(f"\n✓ Tailored cover letters saved to: {base_dir}")
    print(f"  Each letter is customized to the specific job posting")
    print(f"  Organized by tier: EXCELLENT, STRONG, GOOD, POSSIBLE")

if __name__ == "__main__":
    generate_all_tailored_letters()
